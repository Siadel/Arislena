
# 프로그램 구성에 필요한 외부 모듈 가져오기
import discord
import openpyxl
import datetime
import random
import sys
import time
# 프로그램 구성에 필요한 내부 모듈 가져오기
import Ari_data as data
import ipa

# 로딩 시간 재는 용도
time_start = time.time()

# 멤버 정보 획득 권한을 얻는 용도
intents = discord.Intents.default()
intents.members = True

# 권한을 얻은 클라이언트 변수화
ari = discord.Client(intents=intents)

# 유저 명령어 사용 여부를 정하는 용도
play_stopped = False

# system에 있는 데이터 모두를 로드하기
path_users = '/storage/emulated/0/Download/Arislena-master/users/'
filename_system = 'system.xlsx'
system = openpyxl.load_workbook(filename_system)

users_dict = dict()

# 시스템 클래스들








# ./users에 있는 파일 중 디스코드에 가입돼 있는 유저 파일만 로드하기

# 기본 함수


def make_users():
    # 유저별로 
    copies = ['outputs']
    
    
    for guild in ari.guilds:
        for member in guild.members:
            filename = f'{path_users}{member}.xlsx'
            try:
                openpyxl.load_workbook(filename)
                
                
            except FileNotFoundError:
                newfile = openpyxl.Workbook()
                for copi in copies:
                    newfile.create_sheet(copi)
        
                    origin = system[copi]['A']
                    ln = len(origin)
                    
                    for row, n in zip(newfile[copi].iter_rows(max_row=ln, max_col=1), range(ln)):
                        row[0].value = origin[n].value
                    
                    newfile.save(filename)
                    
            users_dict.setdefault(member, openpyxl.load_workbook(filename))



def find_cmd(inpt):
    # 
    abbr = system['abbr']
    
    for row in abbr.iter_rows(values_only=True):
        if inpt in row:
            return row[0]
    
    return

def check_cmds():
    s_cmd = system['cmd']
    cmds = set(cmd_dict.keys())
    system_cmds = set(tuple(s_cmd.iter_cols(values_only=True, max_col = 1))[0])
    remain = cmds - system_cmds
    for re in remain:
        s_cmd.append([re])
    system.save(filename_system)
  
def get_current_time():
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def solve(a, b, c, d, e, f):
	numer_f = e * c - b * f
	numer_m = a * f - d * c
	dnomi = a * e - b * d
	if dnomi == 0:
		raise ValueError()
	return numer_f / dnomi, numer_m / dnomi




def check_permission(roles: discord.Member.roles):
    roles_list = []
    
    for role in roles:
        roles_list.append(role.name)
    
    if '기술자' in roles_list:
        return '기술자'
    elif '주인' in roles_list:
        return '주인'
    elif len(roles_list) == 1:
        return '모두'


def get_channel_by_name(name):
    for ch in ari.get_all_channels():
        if ch.name == name and ch.type == discord.ChannelType.text:
            return ch


def parse(message):
    # !로 명령어를 선언하고, 공백과 따옴표로 명령어 나누기
    message.strip()
    
    if message.startswith('!'):
        message = message[1:]
        
        if len(message) == 0:
            return None
        if '!' in message:
            return None
        message.strip()
        # 따옴표를 사용하려면, 반드시 따옴표 안에 공백을 넣어야 함.

        # 먼저 메세지(str)를 공백을 기준으로 나눈다.
        msg = message.split()
        li = []
        # 나눈 메세지(list)를 스캔해 따옴표가 있는 메세지의 위치를
        # 새로운 리스트에 저장한다.
        for i in range(len(msg)):
            if '"' in msg[i] or "'" in msg[i]:
                li.append(i)
        li.reverse()  # 리스트를 전환한다. (역순 계산을 해야 하기 때문)

        # 따옴표가 있는 메세지의 시작점과 끝 점을 받아 공백으로 잇는다.
        # 이은 메세지를 기존의 메세지와 교체한다.
        for i in range(0, len(li), 2):
            string = ' '.join(msg[li[i + 1]:li[i] + 1])
            del msg[li[i + 1]:li[i] + 1]
            msg.insert(li[i + 1], string)

        # 양쪽의 따옴표를 제거한다.
        # 따옴표가 중간에 존재하는 경우 따옴표를 기준으로 다시 분리한다.
        for i in msg:
            ind = msg.index(i)
            if i.startswith('"') or i.startswith("'"):
                string = i[1:]
                del msg[ind]
                msg.insert(ind, string)
            i = msg[ind]
            if i.endswith('"') or i.endswith("'"):
                string = i[:-1]
                del msg[ind]
                msg.insert(ind, string)
            i = msg[ind]
            if '"' in i:
                nw = i.split('"')
                del msg[ind]
                msg.insert(ind, nw[0])
                msg.insert(ind + 1, nw[1])
            elif "'" in i:
                nw = i.split("'")
                del msg[ind]
                msg.insert(ind, nw[0])
                msg.insert(ind + 1, nw[1])
        
        return msg

    else:
        return None



# 예외 클래스
class ArgsError(Exception):
    # 명령어 자릿수 예외
    def __init__(self, author, args_needed, command, *extra):
        
        if len(extra) != 0:
            ext = ''
        
            if extra[0] == '+':
                ext = '이상이'
            elif extra[0] == '-':
                ext = '이하'

            super().__init__(Outputs('ArgsError_aboveless', author).output.format(args_needed, ext, command))
            
        else:
            super().__init__(Outputs('ArgsError_exact', author).output.format(args_needed, command))


class Permission(Exception):
    # 권한 예외
    def __init__(self, perm_needed, command):
        super().__init__(f'방금 명령어는 **{perm_needed}**만 쓸 수 있어.\n"!아리 {command}" 명령어로 자세히 알아보는 건 어때?')


class DifferentTypeError(Exception):
    # 데이터 종류가 다른 예외
    def __init__(self):
        super().__init__('적당한 값을 안 넣었어.')


class DifferentNameError(Exception):
    # 검색한 나라나 지역이 존재하지 않음을 알리는 오류
    def __init__(self):
        super().__init__('찾고자 하는 건 없어. 오타가 아닐까?')


class MinusError(Exception):
    # 음수 예외
    def __init__(self):
        super().__init__('음수 값을 넣을 순 없어.')


class SumError(Exception):
    # 합 예외
    def __init__(self, string, shouldbe, howmuch):
        if howmuch > shouldbe:
            super().__init__(f'{string} 합이 {shouldbe}이 돼야 해.\n{howmuch - shouldbe}만큼 빼면 되겠다.')
        if howmuch < shouldbe:
            super().__init__(f'{string} 합이 {shouldbe}이 돼야 해.\n{shouldbe - howmuch}만큼 더하면 되겠다.')


class DupliError(Exception):
    def __init__(self):
        super().__init__('이미 참가했어.')


class UnitError(Exception):
    # 구매 시 100단위가 아닌 예외
    def __init__(self):
        super().__init__('꼭 100단위로 써 줘.')


class StopError(Exception):
    # 명령어 작동을 취소한 예외
    def __init__(self):
        super().__init__('하던 작업을 취소했어.')


class CmdNotFound(Exception):
    def __init__(self, author, p_cmd):
        outputsid = 'exceCmdNotFound'
        super().__init__(Outputs(outputsid, author).output.format(p_cmd))


# 아웃풋 클래스
class Outputs:
    def __init__(self, outputs_id, author):
        self.id = outputs_id
        self.author = author
        self.output = self.find_output()
        
    def find_output(self):
        systemoutput = system['output']
        useroutput = users_dict[self.author]['output']
        
        id_col = tuple(systemoutput.iter_cols(values_only=True, max_col=1))[0]
        
        output_dict = dict()
        
        for d in id_col:
            if self.id in d:
                idx = id_col.index(d)+1
                file = ''
            
                if data.indata(useroutput.cell(row=idx, column=2).value):
                    file = useroutput
                else:
                    file = systemoutput
                
                temp_output = list(file[idx])
                del temp_output[0]
                output = []
                for t in temp_output:
                    if data.indata(t.value):
                        output.append(t.value)
                        
                output_dict.setdefault(d, random.choice(output))
            
        if len(output_dict) == 1:
            return list(output_dict.values())[0]
        elif len(output_dict) > 1:
            return output_dict
        
        return


# 상위 명령어 클래스
class Commands:
    
    calls = dict()
    abbrs = []
    
    # 명령어 전반에 대한 가이드 제공
    def __init__(self, norm, category, permission, args_needed, *abbreviations):
        # 명령어의 요소
        self.norm = norm
        self.cate = category
        self.perm = permission
        self.args = args_needed
        self.abbr = abbreviations
        
        Commands.calls.setdefault(norm, abbreviations)
        Commands.abbrs.extend(list(abbreviations))

    async def check(self, message: discord.Message, parsed, permission, args):
        # 기본적인 형식을 검사하는 함수
        try:
            # category
                  
            if self.cate == '행동':
            	pass
            	
            # permission
            
            if not (permission == self.perm or self.perm == '모두' or permission == '기술자'):
                # 명령어 주체가 기술자거나, 권한을 맞추거나, 모두 권한 명령어여야 함
                raise Permission(self.perm, parsed[0])
            	
            # args
            auth = message.author
            
            if '+' in self.args or '-' in self.args:
                s = self.args[:-1]
                if '+' in self.args and args < int(s):
                    raise ArgsError(auth, s, parsed[0], '+')
                if '-' in self.args and args > int(s):
                    raise ArgsError(auth, s, parsed[0], '-')
            elif args != int(self.args):
            	raise ArgsError(auth, self.args, parsed[0])
            
            await self.execute(message, parsed, permission, args)
        except ArgsError as Args:
            await message.channel.send(Args)
        except Permission as Permi:
            await message.channel.send(Permi)
        except NotImplementedError:
            await message.channel.send('그건 아직 구현이 안 됐어.')

    async def execute(self, message: discord.Message, parsed, permission, args):
        # 실행 부분
        raise NotImplementedError()
    
    


# 하위 명령어 클래스 모음
# class (명령어영문명)
# super().__init__((명령어 이름), (명령어 사용 가능 권한), (명령어 자릿수), (명령어 설명), (명령어 세부 설명))

# 기술자
class Exit(Commands):
    def __init__(self):
        super().__init__('종료', '일반', '기술자', '1')
    
    async def execute(self, message, parsed, permission, args):
        
        now = get_current_time()
        channel = get_channel_by_name('실행_종료')
        await channel.send(Outputs('cmdExit_notice', message.author).output.format(now))
        
        sys.exit(1)


class Eolmana(Commands):
    def __init__(self):
        super().__init__('얼마나지났어', '일반', '기술자', '1')

    async def execute(self, message, parsed, permission, args):
        today = datetime.datetime.today()
        y = today.year
        m = today.month
        d = today.day
        howmuch = datetime.datetime(y, m, d) - datetime.datetime(2020, 3, 11)
        await message.channel.send(f'오늘은 개발 시작일부터 **{howmuch.days}**일이 지났어.')


class ConvertSentence(Commands):
	def __init__(self):
		super().__init__('이파', '일반', '기술자', '1+')
		
	async def execute(self, message, parsed, permission, args):
		IPA = []
		HAN = []
		
		for i in range(args-1):
			ipa.exe(parsed[i+1])
			IPA.append(ipa.dresult['ipa_convert'])
			HAN.append(ipa.dresult['meaning_hanja'])
		
		for l in [IPA, HAN]:
			await message.channel.send('\n'.join(l))


# 기술자(기능)
class Stop(Commands):
    def __init__(self):
        super().__init__('중지', '일반', '기술자', '1')

    async def execute(self, message, parsed, permission, args):
        global play_stopped
        if not play_stopped:
            play_stopped = True
            await message.channel.send('타임! 지금부턴 기술자만 명령할 수 있어.')
        elif play_stopped:
            await message.channel.send('이미 중지했어.')


class Restart(Commands):
    def __init__(self):
        super().__init__('재개', '일반', '기술자', '1')

    async def execute(self, message, parsed, permission, args):
        global play_stopped
        if play_stopped:
            play_stopped = False
            await message.channel.send('타임 끝! 이젠 모두 명령할 수 있어.')
        elif not play_stopped:
            await message.channel.send('이미 재개했어.')


class Cycle(Commands):
	def __init__(self):
		super().__init__('새날', '일반', '기술자', '1+')

	async def execute(self, message, parsed, permission, args):
		pass


# 주인
class Settledown(Commands):
	def __init__(self):
		super().__init__('정착', '일반', '주인', '1+')

	async def execute(self, message, parsed, permission, args):
		pass


# 모두
class Ari(Commands):
    def __init__(self):
        super().__init__('아리', '일반', '모두', '1+')

    async def execute(self, message, parsed, permission, args):
        auth = message.author

        if args == 1:
            
            cmdname = 'cmdAri'
            outputs = Outputs(cmdname, auth).output
            print(outputs)
            
            howto_li = []
            perm_li = []
            desc_li = []
            
            
            embs = (len(cmd_dict) - 1) // 9 + 1
            cur = 0
            for v in cmd_dict.values():
                name = v.__name__
                cmd_name_ = f'cmd{name}_'
                
                while cur < 9:
                    if cur == 0:
                        emb = discord.Embed(title = outputs[f'{cmdname}_emb_cmdlist_title'].format())
                    
                    
                    
                    cur += 1
                
            
            li = list(cmd_dict.keys())
            li_perm = []
            li_desc = []
            for i in li:
                li_perm.append(cmd_dict[i].perm)
                li_desc.append('> '+Outputs(cmd_dict[i].desc, auth))
            n = (len(li) - 1) // 9 + 1
            for i in range(n):
                emb = discord.Embed(title=outputs['cmdAri_emb_cmdlist_title'].format(i + 1, n))
                if len(li) > 9:
                    for j in range(9):
                        emb.add_field(name=li[j], value=li_desc[j] + "\n" + '권한: ' + li_perm[j])
                    await message.channel.send(embed=emb)
                    del li[0:9]
                    del li_desc[0:9]
                    del li_perm[0:9]
                elif len(li) <= 9:
                    for j in range(len(li)):
                        emb.add_field(name=li[j], value=li_desc[j] + "\n" + '권한: ' + li_perm[j])
                    await message.channel.send(embed=emb)
        else:
            cmd_li = list(cmd_dict.keys())
            subcmd = ['안녕', '잘가', '알려줘']
            hello = [outputs['cmdAri_hello_nonformat'], outputs['cmdAri_hello_format'].format(auth.name)]
            bye = [outputs['cmdAri_bye_nonformat'], outputs['cmdAri_bye_format'].format(auth.name)]
            ipt = parsed[1]

            if ipt in cmd_li:
                emb = discord.Embed(title=outputs['cmdAri_cmd_emb_title'].format(ipt))
                
                for row in system['cmd'].iter_rows(values_only=True):
                    if ipt in row:
                        abbr = list(row)
                        del abbr[0]
                
                emb.description = ('> '+Outputs(cmd_dict[ipt].desc, auth).output + '\n' + Outputs(cmd_dict[ipt].xpln, auth).output.format(abbr))
                
                await message.channel.send(embed=emb)
            elif ipt in subcmd:
                if ipt == subcmd[0]:
                    await message.channel.send(random.choice(hello))
                elif ipt == subcmd[1]:
                    await message.channel.send(random.choice(bye))
                elif ipt == subcmd[2]:
                    emb = discord.Embed(title=Outputs('cmdAri_help_emb_title', auth))
                    emb.description =  Outputs('cmdAri_help_emb_desc', auth)
                    emb.add_field(name= Outputs('cmdAri_help_emb_field_name', auth), value= Outputs('cmdAri_help_emb_field_value', auth), inline=False)
                    await message.channel.send(embed=emb)


class Signin(Commands):
	def __init__(self):
		super().__init__('참가', '일반', '모두', '2')



# 커맨드 사전
# 커맨드 이름 : 실행
cmd_dict = {Commands.norm: Commands for Commands in [
    # 테스트 명령어(기술자)
    Eolmana(), ConvertSentence(), Exit(),
    # 기술자 명령어
    Stop(), Restart(),
    Cycle(),
    # 주인 명령어
    # 모두 명령어
    Ari(), Signin()
]}



@ari.event
async def on_ready():
    # 프로그램을 켰을 때 실행하는 부분
    
    print(discord.__version__)
    
    make_users()
    check_cmds()
    
    game = discord.Game("아리슬레나의 신 노릇")
    await ari.change_presence(status=discord.Status.online, activity=game)
    
    time_end = time.time()
    time_spn = time_end - time_start
    print(round(time_spn, 3))
    


@ari.event
async def on_member_update(before, after):
    if before.display_name != afterandom.display_name:
        for m in ari.get_all_members():
            if m.display_name == afterandom.display_name and m != after:
                await afterandom.create_dm()
                await afterandom.dm_channel.send("아리슬레나 내부에 중복되는 닉네임이 있습니다. 닉네임 변경이 취소됩니다.")
                await afterandom.edit(nick=before.display_name)

@ari.event
async def on_message(message):
    try:
        global play_stopped
        
        p = parse(message.content)
        auth = message.author
        current_time = get_current_time()
        
        # 유저가 치는 채팅을 인식해서 나눠 p에 할당하는데,
        # 형식상 알맞지 않은 채팅(!로 시작하지 않음)이면 None을 반환한다.
    
        if p is not None: # 명령어가 형식상 적절할 경우
        
            first = p[0]
            
            if not first in Commands.calls.keys():
                
                if not first in Commands.abbrs:
                    
                    raise CmdNotFound(message.author, first)
                
                else:
                    
                    for k, v in Commands.calls.items():
                        
                        if first in v:
                            
                            first = k
        
            if play_stopped and not check_permission(message.author) == '기술자':
                # '중지'되어 있으면, 기술자가 아닌 사람이 명령어를 치면 로그를 콘솔에 출력하고 await 이하를 디코에 출력한다.
                print(current_time, auth, check_permission(message.author), message.content, "CANCELED!", sep="\t")
                await message.channel.send('지금은 기술자만 명령할 수 있어.')
                return
            else:
                # 명령어 상위 클래스인 Commands의 check을 통해 유효성을 검사하고 명령어를 실행한다.
                await cmd_dict[first].check(message, p, check_permission(message.authorandom.roles), len(p))
                print(current_time, auth, check_permission(message.authorandom.roles), len(p), message.content, sep="\t") # 로그 콘솔 출력
                
    except CmdNotFound as cnf: # 명령어가 옳지 않으면 오류 메세지를 출력한다.
        await message.channel.send(cnf)

token = data.token.token

ari.run(token)
